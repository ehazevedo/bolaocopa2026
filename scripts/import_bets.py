#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_GLOB = "*Apostas fase grupos*.xlsx"

ALIASES = {
    "Bosnia": "Bósnia",
    "Bósnia": "Bósnia",
    "Coréia": "Coreia do Sul",
    "Costa do Marfin": "Costa do Marfim",
    "Egiito": "Egito",
    "Egito": "Egito",
    "Austria": "Áustria",
    "Holanda": "Países Baixos",
    "Congo": "RD Congo",
    "Rep Tcheca": "Rep. Tcheca",
}


def clean_name(value):
    if value is None:
        return ""
    text = str(value).strip()
    return ALIASES.get(text, text)


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_only).strip("-").lower()
    return slug or "participante"


def as_score(value) -> int:
    if value is None or value == "":
        return 0
    return int(value)


def participant_name(path: Path, worksheet) -> str:
    cell_name = worksheet["I2"].value
    if cell_name:
        return str(cell_name).strip()
    suffix = "_Apostas fase grupos"
    stem = path.stem
    return stem.replace(suffix, "").strip() or stem


def extract_workbook(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    matches = []
    bets = []
    current_group = ""

    for row in worksheet.iter_rows(min_row=4, max_row=120, values_only=True):
        group_label, match_no, date_value, team_1, goals_1, _, goals_2, team_2 = row[:8]
        if group_label:
            current_group = str(group_label).replace("Grupo", "").strip()
        if not match_no or not team_1 or not team_2:
            continue

        match_id = int(match_no)
        date_iso = ""
        if isinstance(date_value, datetime):
            date_iso = date_value.date().isoformat()

        matches.append(
            {
                "id": match_id,
                "phase": "Fase de Grupos",
                "group": current_group,
                "date": date_iso,
                "team1": clean_name(team_1),
                "team2": clean_name(team_2),
            }
        )
        bets.append(
            {
                "matchId": match_id,
                "g1": as_score(goals_1),
                "g2": as_score(goals_2),
            }
        )

    name = participant_name(path, worksheet)
    return {
        "participant": {
            "id": slugify(name),
            "name": name,
            "file": str(path.name),
            "bets": bets,
        },
        "matches": matches,
    }


def dedupe_id(participant, existing):
    base = participant["id"]
    candidate = base
    counter = 2
    while candidate in existing:
        candidate = f"{base}-{counter}"
        counter += 1
    participant["id"] = candidate


def build_data(input_dir: Path):
    files = sorted(input_dir.glob(WORKBOOK_GLOB))
    if not files:
        raise SystemExit(f"Nenhum arquivo encontrado em {input_dir} com padrão {WORKBOOK_GLOB!r}.")

    participants = []
    seen_ids = set()
    canonical_matches = None
    warnings = []

    for path in files:
        extracted = extract_workbook(path)
        participant = extracted["participant"]
        dedupe_id(participant, seen_ids)
        seen_ids.add(participant["id"])
        participants.append(participant)

        if canonical_matches is None:
            canonical_matches = extracted["matches"]
        else:
            current = [(m["id"], m["team1"], m["team2"]) for m in extracted["matches"]]
            canonical = [(m["id"], m["team1"], m["team2"]) for m in canonical_matches]
            if current != canonical:
                warnings.append(f"{path.name}: lista de jogos difere da primeira planilha importada.")

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFolder": str(input_dir),
        "rules": {
            "simpleResultPoints": 2,
            "exactScoreBonus": 3,
            "maxPerMatch": 5,
            "stageWeights": {
                "Fase de Grupos": 35,
                "Rodada de 32": 25,
                "Oitavas à Final": 40,
            },
        },
        "matches": canonical_matches or [],
        "participants": participants,
        "warnings": warnings,
    }


def main():
    parser = argparse.ArgumentParser(description="Importa apostas do bolão para o dashboard HTML.")
    parser.add_argument(
        "--input",
        default="apostas",
        help="Pasta com arquivos *_Apostas fase grupos.xlsx.",
    )
    parser.add_argument(
        "--output",
        default="data/bolao-data.js",
        help="Arquivo JS gerado para uso pelo dashboard.",
    )
    args = parser.parse_args()

    data = build_data(Path(args.input).expanduser().resolve())
    output = Path(args.output).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    js = "window.BOLAO_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
    output.write_text(js, encoding="utf-8")
    print(f"Importados {len(data['participants'])} participante(s), {len(data['matches'])} jogo(s).")
    print(f"Dados salvos em {output}")
    if data["warnings"]:
        print("Avisos:")
        for warning in data["warnings"]:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
